#!/usr/bin/env python3
"""
Parses data from CSV, JSON, or Excel and formats it for use in Highcharts.

Automatically detects format, encoding, and data structure.
Output: JSON ready to be embedded in Highcharts options.

Usage:
    python parse_data.py <file> [--format categories|timeseries|xy|pie]
    python parse_data.py data.csv --sheet "Sheet1" --encoding utf-8
    python parse_data.py data.json --output formatted.json

Output:
    JSON with: { categories, series, metadata }
"""

import sys
import json
import argparse
from pathlib import Path


def detect_encoding(filepath: str) -> str:
    """Attempts to detect the file encoding."""
    encodings = ['utf-8', 'utf-8-sig', 'latin1', 'iso-8859-1', 'cp1252']
    for enc in encodings:
        try:
            with open(filepath, 'r', encoding=enc) as f:
                f.read(1000)
            return enc
        except (UnicodeDecodeError, UnicodeError):
            continue
    return 'utf-8'


def parse_number(value: str) -> float | None:
    """Converts a string to a number, handling BR and US formats."""
    if not value or not isinstance(value, str):
        return value if isinstance(value, (int, float)) else None
    value = value.strip().replace(' ', '')
    # BR format: 1.234,56
    if ',' in value and '.' in value and value.rindex(',') > value.rindex('.'):
        value = value.replace('.', '').replace(',', '.')
    # BR format without thousands separator: 123,45
    elif ',' in value and '.' not in value:
        value = value.replace(',', '.')
    # Remove currency symbols
    for symbol in ['R$', '$', '€', '£', '%']:
        value = value.replace(symbol, '')
    try:
        return float(value)
    except ValueError:
        return None


def parse_csv(filepath: str, encoding: str = 'utf-8', delimiter: str = None) -> dict:
    """Parses a CSV into Highcharts format."""
    import csv

    with open(filepath, 'r', encoding=encoding) as f:
        content = f.read()

    # Detect delimiter
    if delimiter is None:
        sniffer = csv.Sniffer()
        try:
            dialect = sniffer.sniff(content[:2000])
            delimiter = dialect.delimiter
        except csv.Error:
            delimiter = ',' if ',' in content else ';' if ';' in content else '\t'

    lines = content.strip().split('\n')
    reader = csv.reader(lines, delimiter=delimiter)
    rows = list(reader)

    if len(rows) < 2:
        return {"error": "File has fewer than 2 lines"}

    headers = [h.strip() for h in rows[0]]
    data_rows = rows[1:]

    # First column = categories, rest = series
    categories = [row[0].strip() for row in data_rows if row]
    series = []
    for col_idx in range(1, len(headers)):
        values = []
        for row in data_rows:
            if col_idx < len(row):
                val = parse_number(row[col_idx])
                values.append(val if val is not None else 0)
            else:
                values.append(0)
        series.append({
            "name": headers[col_idx],
            "data": values
        })

    return {
        "categories": categories,
        "series": series,
        "metadata": {
            "rows": len(data_rows),
            "columns": len(headers),
            "headers": headers,
            "delimiter": delimiter,
            "encoding": encoding
        }
    }


def parse_json_data(filepath: str) -> dict:
    """Parses JSON into Highcharts format."""
    with open(filepath, 'r', encoding='utf-8') as f:
        data = json.load(f)

    # If already in Highcharts format, return directly
    if isinstance(data, dict) and 'series' in data:
        return data

    # If it's an array of objects [{x: ..., y: ...}]
    if isinstance(data, list) and len(data) > 0 and isinstance(data[0], dict):
        keys = list(data[0].keys())
        category_key = keys[0]
        categories = [str(item.get(category_key, '')) for item in data]
        series = []
        for key in keys[1:]:
            values = [parse_number(str(item.get(key, 0))) or 0 for item in data]
            series.append({"name": key, "data": values})
        return {
            "categories": categories,
            "series": series,
            "metadata": {"rows": len(data), "keys": keys, "format": "array_of_objects"}
        }

    # If it's an array of arrays [[cat, v1, v2], ...]
    if isinstance(data, list) and len(data) > 0 and isinstance(data[0], list):
        categories = [str(row[0]) for row in data[1:]]
        headers = data[0]
        series = []
        for col_idx in range(1, len(headers)):
            values = [parse_number(str(row[col_idx])) or 0
                      for row in data[1:] if col_idx < len(row)]
            series.append({"name": str(headers[col_idx]), "data": values})
        return {
            "categories": categories,
            "series": series,
            "metadata": {"rows": len(data) - 1, "format": "array_of_arrays"}
        }

    return {"error": "Unrecognized JSON format", "raw": data}


def parse_excel(filepath: str, sheet: str = None) -> dict:
    """Parses Excel into Highcharts format."""
    from openpyxl import load_workbook

    wb = load_workbook(filepath, read_only=True, data_only=True)

    if sheet:
        ws = wb[sheet]
    else:
        ws = wb.active

    rows = list(ws.iter_rows(values_only=True))
    if len(rows) < 2:
        return {"error": "Spreadsheet has fewer than 2 rows"}

    headers = [str(h).strip() if h else f"Col_{i}" for i, h in enumerate(rows[0])]
    data_rows = rows[1:]

    categories = [str(row[0]).strip() if row[0] else '' for row in data_rows]
    series = []
    for col_idx in range(1, len(headers)):
        values = []
        for row in data_rows:
            val = row[col_idx] if col_idx < len(row) else 0
            if isinstance(val, (int, float)):
                values.append(val)
            else:
                parsed = parse_number(str(val)) if val else 0
                values.append(parsed or 0)
        series.append({"name": headers[col_idx], "data": values})

    return {
        "categories": categories,
        "series": series,
        "metadata": {
            "rows": len(data_rows),
            "columns": len(headers),
            "headers": headers,
            "sheet": ws.title,
            "sheets_available": wb.sheetnames
        }
    }


def main():
    parser = argparse.ArgumentParser(description="Parses data into Highcharts format")
    parser.add_argument("filepath", help="Path to the data file")
    parser.add_argument("--encoding", default=None, help="CSV file encoding")
    parser.add_argument("--delimiter", default=None, help="CSV delimiter")
    parser.add_argument("--sheet", default=None, help="Excel sheet name")
    parser.add_argument("--output", "-o", help="Save result to file")
    args = parser.parse_args()

    path = Path(args.filepath)
    if not path.exists():
        print(f"[ERROR] File not found: {path}", file=sys.stderr)
        sys.exit(1)

    ext = path.suffix.lower()

    if ext in ('.csv', '.tsv', '.txt'):
        encoding = args.encoding or detect_encoding(str(path))
        result = parse_csv(str(path), encoding=encoding, delimiter=args.delimiter)
    elif ext == '.json':
        result = parse_json_data(str(path))
    elif ext in ('.xlsx', '.xls'):
        result = parse_excel(str(path), sheet=args.sheet)
    else:
        print(f"[ERROR] Unsupported format: {ext}", file=sys.stderr)
        sys.exit(1)

    if "error" in result:
        print(f"[ERROR] {result['error']}", file=sys.stderr)
        sys.exit(1)

    meta = result.get("metadata", {})
    print(f"[INFO] Rows: {meta.get('rows', '?')}, "
          f"Columns: {meta.get('columns', len(result.get('series', [])) + 1)}", file=sys.stderr)
    if 'series' in result:
        for s in result['series']:
            print(f"[INFO]   Series '{s['name']}': {len(s['data'])} points", file=sys.stderr)

    output = json.dumps(result, ensure_ascii=False, indent=2)

    if args.output:
        with open(args.output, 'w', encoding='utf-8') as f:
            f.write(output)
        print(f"[INFO] Saved to: {args.output}", file=sys.stderr)
    else:
        print(output)


if __name__ == "__main__":
    main()
